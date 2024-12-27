import sys
import os
import pandas as pd
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QPushButton,
    QLabel,
    QFileDialog,
    QVBoxLayout,
    QHBoxLayout,
    QLineEdit,
    QMessageBox,
)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelAutomationApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("중앙도서관 엑셀 노예")
        self.setGeometry(100, 100, 700, 400)
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        # 엑셀 파일 선택
        excel_layout = QHBoxLayout()
        self.excel_label = QLabel("작업할 엑셀 파일:")
        self.excel_path = QLineEdit()
        self.excel_path.setReadOnly(True)
        self.excel_button = QPushButton("파일 선택")
        self.excel_button.clicked.connect(self.select_worker_scan_file)
        excel_layout.addWidget(self.excel_label)
        excel_layout.addWidget(self.excel_path)
        excel_layout.addWidget(self.excel_button)
        layout.addLayout(excel_layout)

        # .txt 파일 생성 버튼
        self.txt_button = QPushButton(".txt 파일 생성")
        self.txt_button.clicked.connect(self.create_txt_file)
        layout.addWidget(self.txt_button)

        # 출력용 엑셀 파일 선택
        output_layout = QHBoxLayout()
        self.output_label = QLabel("출력용 엑셀 파일 선택:")
        self.output_path = QLineEdit()
        self.output_path.setReadOnly(True)
        self.output_button = QPushButton("파일 선택")
        self.output_button.clicked.connect(self.select_output_worker_scan_file)
        output_layout.addWidget(self.output_label)
        output_layout.addWidget(self.output_path)
        output_layout.addWidget(self.output_button)
        layout.addLayout(output_layout)

        # 구분 텍스트 입력
        separator_layout = QHBoxLayout()
        self.separator_label = QLabel("구분 텍스트:")
        self.separator_input = QLineEdit()
        separator_layout.addWidget(self.separator_label)
        separator_layout.addWidget(self.separator_input)
        layout.addLayout(separator_layout)

        # 출력 엑셀 파일 생성 버튼
        self.generate_output_button = QPushButton("출력용 엑셀 파일 생성")
        self.generate_output_button.clicked.connect(self.generate_tulip_file)
        layout.addWidget(self.generate_output_button)

        # 상태 표시
        self.status_label = QLabel("상태: 대기 중")
        self.status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.status_label)

        self.setLayout(layout)

    def select_worker_scan_file(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "작업할 데이터 파일 선택",
            "",
            "Excel Files (*.xlsx *.xls)",
            options=options,
        )
        if file_name:
            self.excel_path.setText(file_name)

    def select_output_worker_scan_file(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "출력용 엑셀 파일 선택",
            "",
            "Excel Files (*.xlsx *.xls)",
            options=options,
        )
        if file_name:
            self.output_path.setText(file_name)

    def create_txt_file(self):
        worker_scan_file = self.excel_path.text()
        if not worker_scan_file:
            QMessageBox.warning(self, "경고", "작업할 엑셀 파일을 선택해주세요.")
            return
        try:
            # 파일 확장자에 따라 적절한 엔진을 선택
            file_extension = os.path.splitext(worker_scan_file)[1].lower()
            if file_extension == ".xlsx":
                engine = "openpyxl"
            elif file_extension == ".xls":
                engine = "xlrd"
            else:
                raise ValueError(
                    "지원되지 않는 파일 형식입니다. .xlsx 또는 .xls 파일을 선택해주세요."
                )

            df = pd.read_excel(worker_scan_file, header=None, engine=engine)

            # df를 불러와 행 공백 탐지: 공백 행, 공백 행 + 1의 행 번호를 저장 후 drop
            blank_rows = df.index[df.isnull().all(axis=1)]
            rows_to_drop = blank_rows.union(blank_rows + 1)
            df.drop(rows_to_drop, inplace=True, errors="ignore")
            df = df.drop([0])

            # 엑셀파일의 행을 읽어오고, 공백 행을 기준으로 파티션을 나누어 .txt 파일로 저장
            partitions = []
            current_partition = []
            for _, row in df.iterrows():
                if row.isnull().all():
                    if current_partition:
                        partitions.append(current_partition)
                        current_partition = []
                else:
                    current_partition.append(row.tolist())
            if current_partition:
                partitions.append(current_partition)

            # 행 개수를 카운트해 출력
            cleaned_data = [line for partition in partitions for line in partition]
            cleaned_df = pd.DataFrame(cleaned_data)
            count = len(cleaned_df)
            txt_file = os.path.splitext(worker_scan_file)[0] + "_text.txt"
            cleaned_df.to_csv(txt_file, index=False, header=False, encoding="utf-8")

            QMessageBox.information(
                self, "성공", f".txt 파일이 성공적으로 생성되었습니다.\n개수: {count}"
            )
        except Exception as e:
            QMessageBox.critical(
                self, "에러", f".txt 파일 생성 중 오류가 발생했습니다:\n{str(e)}"
            )

    def generate_tulip_file(self):
        worker_scan_file = self.excel_path.text()  # 작업자 엑셀파일
        tulip_file = self.output_path.text()  # 튤립 export 엑셀파일
        separator_text = self.separator_input.text()

        if not worker_scan_file or not tulip_file:
            QMessageBox.warning(
                self, "경고", "작업할 엑셀 파일과 출력용 엑셀 파일을 선택해주세요."
            )
            return

        if not separator_text:
            QMessageBox.warning(self, "경고", "구분 텍스트를 입력해주세요.")
            return

        try:
            self.status_label.setText("상태: 처리 중...")
            QApplication.processEvents()

            # Step 1: Load and process worker_scan_file to identify partitions
            worker_df = pd.read_excel(worker_scan_file, header=None)

            # 행 공백 기준으로 데이터 분할
            blank_rows = worker_df.isnull().all(axis=1)

            partitions = []
            current_partition = []
            separator_nums = []

            for index, row in worker_df.iterrows():
                if blank_rows[index]:
                    if current_partition:
                        separator_nums.append(current_partition[0][0])
                        partitions.append(current_partition)
                        current_partition = []
                else:
                    current_partition.append(row.tolist())

            # 마지막 파티션 추가
            if current_partition:
                partitions.append(current_partition)
                separator_nums.append(current_partition[0][0])

            # 각 파티션의 행 개수 카운트
            partition_dict = {}
            # key: separator_nums, value: partition의 행 개수
            for i, partition in enumerate(partitions):
                partition_dict[separator_nums[i]] = len(partition) - 1

            # Step 2: Load and process tulip_file
            tulip_df = pd.read_excel(tulip_file, header=None)
            tulip_df = tulip_df.drop([0, 1, 2, 3])
            tulip_df = tulip_df.iloc[:, [0, 1, 2, 4, 5, 10, 11]]

            # 열 이름 지정
            columns = [
                "No.",
                "등록번호",
                "서명",
                "출판사",
                "출판년",
                "소장처",
                "자료실",
            ]

            # Step 3: Prepare tulip_df for processing
            tulip_df = tulip_df.reset_index(drop=True)

            # Step 4: Create the output data with partition separators
            output_data = []
            first_partition_flag = True

            for partition_num, partition_size in partition_dict.items():
                separator_row = [f"{separator_text}-{partition_num}"] + [""] * (
                    len(columns) - 1
                )
                if first_partition_flag:
                    output_data.append(separator_row)
                    first_partition_flag = False
                else:
                    empty_row = [""] * len(columns)
                    output_data.append(empty_row)
                    output_data.append(columns)
                    output_data.append(separator_row)

                # 이제 tulip_df에서 partition_size만큼 뽑기
                data = tulip_df.iloc[:partition_size].copy()  # partition_size 행만큼
                tulip_df.drop(
                    tulip_df.index[:partition_size], inplace=True
                )  # 사용한 만큼 drop
                tulip_df.reset_index(drop=True, inplace=True)

                # reset index and update No. column
                data.reset_index(drop=True, inplace=True)
                data.index = data.index + 1
                data.iloc[:, 0] = data.index
                # 뽑은 데이터 이어 붙이기
                output_data.extend(data.values.tolist())

            # Step 5: Convert the output data to a DataFrame
            output_df = pd.DataFrame(output_data, columns=columns)

            # Load the workbook
            wb = load_workbook(worker_scan_file)

            # Check if '출력용' sheet exists
            sheet_name = "출력용"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Clear existing contents by removing the sheet and creating a new one
                wb.remove(ws)
                ws = wb.create_sheet(title=sheet_name)
            else:
                # Add a new sheet
                ws = wb.create_sheet(title=sheet_name)

            # Write the DataFrame to the '출력용' sheet
            for r_idx, row in enumerate(
                dataframe_to_rows(output_df, index=False, header=True), 1
            ):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Save the workbook
            wb.save(worker_scan_file)

            self.status_label.setText("상태: 완료")
            QMessageBox.information(
                self, "성공", f"'{sheet_name}' 시트가 성공적으로 생성되었습니다."
            )
        except FileNotFoundError:
            self.status_label.setText("상태: 오류 발생")
            QMessageBox.critical(self, "에러", "선택한 파일을 찾을 수 없습니다.")
        except pd.errors.EmptyDataError:
            self.status_label.setText("상태: 오류 발생")
            QMessageBox.critical(self, "에러", "선택한 파일에 데이터가 없습니다.")
        except Exception as e:
            self.status_label.setText("상태: 오류 발생")
            QMessageBox.critical(
                self, "에러", f"출력용 엑셀 파일 생성 중 오류가 발생했습니다:\n{str(e)}"
            )


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelAutomationApp()
    window.show()
    sys.exit(app.exec_())
